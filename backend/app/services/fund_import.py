import re
from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.constants import CATEGORY_SET, normalize_category
from app.models.fund import Fund, FundCode
from app.models.tag import FundTag, Tag
from app.models.tier import FundCurrentTier
from app.services.tushare_sync import lookup_fund_basic


_CODE_RE = re.compile(r"^\d{6}$")
_MARKET_RE = re.compile(r"^(OF|SH|SZ)$", re.IGNORECASE)
_TAG_SEP_RE = re.compile(r"[,，;；、]")


def _normalize_market(value) -> str:
    if not value:
        return "OF"
    s = str(value).strip().upper()
    if s in ("OF", "场外"):
        return "OF"
    if s in ("SH", "上海", "沪"):
        return "SH"
    if s in ("SZ", "深圳", "深"):
        return "SZ"
    return s


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _to_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip().replace("-", "").replace("/", "")
    if len(s) == 8 and s.isdigit():
        return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    return None


def _parse_tags(value) -> List[str]:
    if not value:
        return []
    s = str(value).strip()
    if not s:
        return []
    return [t.strip() for t in _TAG_SEP_RE.split(s) if t.strip()]


def _resolve_tags(db: Session, tag_names: List[str]) -> Tuple[List[UUID], List[str]]:
    tag_ids = []
    created = []
    existing_by_name = {t.name: t for t in db.query(Tag).filter(Tag.is_active.is_(True)).all()}
    for name in tag_names:
        tag = existing_by_name.get(name)
        if not tag:
            tag = Tag(name=name, category="策略/主题", is_active=True)
            db.add(tag)
            db.flush()
            existing_by_name[name] = tag
            created.append(name)
        tag_ids.append(tag.id)
    return tag_ids, created


def _lookup_basic(code: str, market: str) -> Optional[Dict]:
    if not settings.TUSHARE_TOKEN:
        return None
    try:
        return lookup_fund_basic(code, market)
    except Exception:
        return None


def import_funds_from_excel(db: Session, content: bytes) -> Dict:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active
    if sheet is None:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": ["Excel 文件没有工作表"]}

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": ["Excel 文件为空"]}

    headers = [str(h).strip() if h else "" for h in rows[0]]
    header_map = {h.lower(): idx for idx, h in enumerate(headers)}

    def col(name_variants: List[str]) -> Optional[int]:
        for variant in name_variants:
            idx = header_map.get(variant.lower())
            if idx is not None:
                return idx
        return None

    code_idx = col(["基金代码", "code", "基金代码code"])
    market_idx = col(["市场", "market"])
    name_idx = col(["基金名称", "名称", "name"])
    category_idx = col(["分类", "类别", "category"])
    risk_idx = col(["风险等级", "风险", "risk_level", "risk"])
    manager_idx = col(["基金经理", "经理", "manager"])
    establish_idx = col(["成立日期", "establish_date", "成立日"])
    reason_idx = col(["入选理由", "reason", "推荐理由"])
    target_idx = col(["目标客户", "适用客群", "target_clients", "客群"])
    tags_idx = col(["标签", "tags", "tag"])

    if code_idx is None:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": ["缺少基金代码列"]}

    created_count = 0
    updated_count = 0
    skipped_count = 0
    errors: List[str] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        code_val = row[code_idx] if code_idx < len(row) else None
        code = _to_str(code_val)
        if not code:
            continue

        if not _CODE_RE.match(code):
            errors.append(f"第 {row_idx} 行基金代码格式错误：{code_val}")
            skipped_count += 1
            continue

        market = _normalize_market(row[market_idx] if market_idx is not None and market_idx < len(row) else None)
        if not _MARKET_RE.match(market):
            errors.append(f"第 {row_idx} 行市场 {market} 不合法，应为 OF/SH/SZ")
            skipped_count += 1
            continue

        name = _to_str(row[name_idx]) if name_idx is not None else None
        category = _to_str(row[category_idx]) if category_idx is not None else None
        risk_level = _to_str(row[risk_idx]) if risk_idx is not None else None
        manager = _to_str(row[manager_idx]) if manager_idx is not None else None
        establish_date = _to_date(row[establish_idx]) if establish_idx is not None else None
        reason = _to_str(row[reason_idx]) if reason_idx is not None else None
        target_clients = _to_str(row[target_idx]) if target_idx is not None else None
        tags = _parse_tags(row[tags_idx]) if tags_idx is not None else []

        # Validate user-provided category against the allowed PRD category set.
        user_category = normalize_category(category) if category else None
        if category and user_category is None:
            errors.append(f"第 {row_idx} 行 {code} 的分类“{category}”不在系统分类中")
            skipped_count += 1
            continue

        # Try to fill missing basic info from Tushare
        if not name or not category or not establish_date:
            basic = _lookup_basic(code, market)
            if basic:
                name = name or basic.get("name")
                category = category or basic.get("fund_type")
                establish_date = establish_date or basic.get("found_date")

        # Normalize final category; fall back to "其他" for unrecognized Tushare types.
        category = normalize_category(category)
        if category is None:
            category = "其他"

        if not name or not category or not risk_level:
            errors.append(f"第 {row_idx} 行 {code} 缺少必填字段（名称、分类、风险等级）")
            skipped_count += 1
            continue

        existing_code = db.query(FundCode).filter(FundCode.code == code).first()
        tag_ids, created_tags = _resolve_tags(db, tags) if tags else ([], [])

        try:
            if existing_code:
                fund = existing_code.fund
                fund.name = name
                fund.category = category
                fund.risk_level = risk_level
                if manager:
                    fund.manager = manager
                if establish_date:
                    fund.establish_date = establish_date
                if reason:
                    fund.reason = reason
                if target_clients:
                    fund.target_clients = target_clients

                if tags:
                    db.query(FundTag).filter(FundTag.fund_id == fund.id).delete(synchronize_session=False)
                    for tag_id in tag_ids:
                        db.add(FundTag(fund_id=fund.id, tag_id=tag_id))

                db.commit()
                updated_count += 1
            else:
                fund = Fund(
                    name=name,
                    category=category,
                    risk_level=risk_level,
                    manager=manager,
                    establish_date=establish_date,
                    reason=reason,
                    target_clients=target_clients,
                )
                db.add(fund)
                db.flush()

                fund_code = FundCode(
                    fund_id=fund.id,
                    code=code,
                    market=market,
                    is_primary=True,
                )
                tier = FundCurrentTier(fund_id=fund.id)
                db.add_all([fund_code, tier])

                for tag_id in tag_ids:
                    db.add(FundTag(fund_id=fund.id, tag_id=tag_id))

                db.commit()
                created_count += 1
        except Exception as exc:
            db.rollback()
            errors.append(f"第 {row_idx} 行 {code} 导入失败：{exc}")
            skipped_count += 1

    return {
        "created": created_count,
        "updated": updated_count,
        "skipped": skipped_count,
        "errors": errors,
    }
